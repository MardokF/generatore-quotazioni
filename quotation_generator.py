import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import ColumnChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
from datetime import datetime
import json

class QuotationGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("🚀 Generatore Quotazioni Progetti v2.0")
        self.root.geometry("1000x700")
        self.root.configure(bg='#f0f0f0')
        
        # Configurazione stili
        self.setup_styles()
        
        # Variabili
        self.selected_architecture = tk.StringVar(value="enterprise")
        self.baseline_count = tk.IntVar(value=3)
        self.project_name = tk.StringVar(value="Nuovo Progetto")
        self.client_name = tk.StringVar(value="")
        self.project_description = tk.StringVar(value="")
        
        # Architetture predefinite
        self.architectures = {
            "web-app": {
                "name": "🌐 Web Application",
                "description": "Frontend, Backend, Database, API, Testing",
                "items": [
                    "Frontend Development", "Backend Development", "Database Design",
                    "API Development", "UI/UX Design", "Testing & QA",
                    "DevOps & Deployment", "Project Management"
                ]
            },
            "mobile-app": {
                "name": "📱 Mobile Application", 
                "description": "iOS, Android, Backend, API, Store Deployment",
                "items": [
                    "iOS Development", "Android Development", "Backend Services",
                    "API Integration", "UI/UX Design", "Testing Mobile",
                    "App Store Deployment", "Push Notifications", "Project Management"
                ]
            },
            "enterprise": {
                "name": "🏢 Enterprise Solution",
                "description": "Microservizi, Integration, Security, Monitoring",
                "items": [
                    "Architecture Design", "Microservices Development", "Integration Layer",
                    "Security Implementation", "Monitoring & Logging", "Data Migration",
                    "Performance Optimization", "Documentation", "Training", "Project Management"
                ]
            },
            "data-platform": {
                "name": "📈 Data Platform",
                "description": "ETL, Analytics, Reporting, ML Pipeline",
                "items": [
                    "Data Ingestion", "ETL Development", "Data Warehouse Design",
                    "Analytics Dashboard", "ML Pipeline", "Data Governance",
                    "Reporting Tools", "Performance Tuning", "Project Management"
                ]
            }
        }
        
        # Baseline data storage
        self.baseline_data = []
        
        # Setup UI
        self.create_ui()
        self.update_baselines()
        
    def setup_styles(self):
        """Configura gli stili dell'interfaccia"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configura colori personalizzati
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), background='#f0f0f0')
        style.configure('Heading.TLabel', font=('Arial', 12, 'bold'), background='#f0f0f0')
        style.configure('Custom.TButton', font=('Arial', 10, 'bold'))
        
    def create_ui(self):
        """Crea l'interfaccia utente completa"""
        
        # Header
        header_frame = tk.Frame(self.root, bg='#2c3e50', height=80)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="🚀 Generatore Quotazioni Progetti", 
                              font=('Arial', 20, 'bold'), fg='white', bg='#2c3e50')
        title_label.pack(pady=20)
        
        # Main container con scrollbar
        main_container = tk.Frame(self.root, bg='#f0f0f0')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Canvas per scrolling
        canvas = tk.Canvas(main_container, bg='#f0f0f0')
        scrollbar = ttk.Scrollbar(main_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Sezioni dell'interfaccia
        self.create_project_info_section(scrollable_frame)
        self.create_baseline_section(scrollable_frame)
        self.create_architecture_section(scrollable_frame)
        self.create_rates_section(scrollable_frame)
        self.create_action_section(scrollable_frame)
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
    def create_project_info_section(self, parent):
        """Sezione informazioni progetto"""
        section_frame = self.create_section_frame(parent, "📋 Informazioni Progetto")
        
        # Grid per i campi
        info_grid = tk.Frame(section_frame, bg='white')
        info_grid.pack(fill='x', padx=20, pady=10)
        
        # Nome progetto
        tk.Label(info_grid, text="Nome Progetto:", font=('Arial', 10, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=5)
        project_entry = tk.Entry(info_grid, textvariable=self.project_name, font=('Arial', 10), width=40)
        project_entry.grid(row=0, column=1, sticky='ew', pady=5, padx=(10, 0))
        
        # Cliente
        tk.Label(info_grid, text="Cliente:", font=('Arial', 10, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=5)
        client_entry = tk.Entry(info_grid, textvariable=self.client_name, font=('Arial', 10), width=40)
        client_entry.grid(row=1, column=1, sticky='ew', pady=5, padx=(10, 0))
        
        # Descrizione
        tk.Label(info_grid, text="Descrizione:", font=('Arial', 10, 'bold'), bg='white').grid(row=2, column=0, sticky='nw', pady=5)
        desc_text = tk.Text(info_grid, height=3, width=40, font=('Arial', 10))
        desc_text.grid(row=2, column=1, sticky='ew', pady=5, padx=(10, 0))
        desc_text.bind('<KeyRelease>', lambda e: self.project_description.set(desc_text.get('1.0', 'end-1c')))
        
        info_grid.columnconfigure(1, weight=1)
        
    def create_baseline_section(self, parent):
        """Sezione configurazione baseline"""
        section_frame = self.create_section_frame(parent, "📊 Configurazione Baseline")
        
        # Controllo numero baseline
        baseline_control = tk.Frame(section_frame, bg='white')
        baseline_control.pack(fill='x', padx=20, pady=10)
        
        tk.Label(baseline_control, text="Numero di Baseline:", font=('Arial', 10, 'bold'), bg='white').pack(side='left')
        baseline_spin = tk.Spinbox(baseline_control, from_=2, to=5, textvariable=self.baseline_count, 
                                  command=self.update_baselines, width=5, font=('Arial', 10))
        baseline_spin.pack(side='left', padx=(10, 0))
        
        # Container per le baseline
        self.baseline_container = tk.Frame(section_frame, bg='white')
        self.baseline_container.pack(fill='x', padx=20, pady=10)
        
    def create_architecture_section(self, parent):
        """Sezione selezione architettura"""
        section_frame = self.create_section_frame(parent, "🏗️ Architettura di Progetto")
        
        arch_grid = tk.Frame(section_frame, bg='white')
        arch_grid.pack(fill='x', padx=20, pady=10)
        
        row = 0
        col = 0
        for arch_key, arch_data in self.architectures.items():
            arch_frame = tk.Frame(arch_grid, bg='#f8f9fa', relief='ridge', bd=2)
            arch_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
            
            radio = tk.Radiobutton(arch_frame, text=arch_data["name"], variable=self.selected_architecture,
                                 value=arch_key, bg='#f8f9fa', font=('Arial', 10, 'bold'),
                                 command=self.on_architecture_change)
            radio.pack(anchor='w', padx=10, pady=5)
            
            desc_label = tk.Label(arch_frame, text=arch_data["description"], bg='#f8f9fa', 
                                 font=('Arial', 9), wraplength=200, justify='left')
            desc_label.pack(anchor='w', padx=10, pady=(0, 10))
            
            col += 1
            if col > 1:
                col = 0
                row += 1
                
        # Configura il grid
        for i in range(2):
            arch_grid.columnconfigure(i, weight=1)
    
    def create_rates_section(self, parent):
        """Sezione configurazione tariffe"""
        section_frame = self.create_section_frame(parent, "💰 Configurazione Tariffe")
        
        rates_frame = tk.Frame(section_frame, bg='white')
        rates_frame.pack(fill='x', padx=20, pady=10)
        
        # Default rates
        self.rates = {
            "Senior Developer": 800,
            "Developer": 600,
            "Junior Developer": 400,
            "Project Manager": 900,
            "Business Analyst": 700,
            "QA Tester": 500
        }
        
        self.rate_vars = {}
        
        # Header
        tk.Label(rates_frame, text="Ruolo", font=('Arial', 10, 'bold'), bg='white', width=20).grid(row=0, column=0, pady=5)
        tk.Label(rates_frame, text="Tariffa/giorno (€)", font=('Arial', 10, 'bold'), bg='white').grid(row=0, column=1, pady=5)
        
        # Rate entries
        for i, (role, rate) in enumerate(self.rates.items(), 1):
            tk.Label(rates_frame, text=role, bg='white', font=('Arial', 10)).grid(row=i, column=0, sticky='w', pady=2)
            
            rate_var = tk.IntVar(value=rate)
            self.rate_vars[role] = rate_var
            rate_entry = tk.Entry(rates_frame, textvariable=rate_var, width=10, font=('Arial', 10))
            rate_entry.grid(row=i, column=1, pady=2, padx=(10, 0))
    
    def create_action_section(self, parent):
        """Sezione azioni"""
        action_frame = tk.Frame(parent, bg='#f0f0f0')
        action_frame.pack(fill='x', pady=20)
        
        # Pulsanti
        button_frame = tk.Frame(action_frame, bg='#f0f0f0')
        button_frame.pack()
        
        generate_btn = tk.Button(button_frame, text="🚀 Genera Template Excel", 
                               command=self.generate_excel, bg='#28a745', fg='white',
                               font=('Arial', 12, 'bold'), padx=20, pady=10)
        generate_btn.pack(side='left', padx=10)
        
        preview_btn = tk.Button(button_frame, text="👁️ Anteprima Dati", 
                              command=self.preview_data, bg='#17a2b8', fg='white',
                              font=('Arial', 12, 'bold'), padx=20, pady=10)
        preview_btn.pack(side='left', padx=10)
        
        save_config_btn = tk.Button(button_frame, text="💾 Salva Configurazione", 
                                   command=self.save_configuration, bg='#ffc107', fg='black',
                                   font=('Arial', 12, 'bold'), padx=20, pady=10)
        save_config_btn.pack(side='left', padx=10)
        
        load_config_btn = tk.Button(button_frame, text="📁 Carica Configurazione", 
                                   command=self.load_configuration, bg='#6c757d', fg='white',
                                   font=('Arial', 12, 'bold'), padx=20, pady=10)
        load_config_btn.pack(side='left', padx=10)
    
    def create_section_frame(self, parent, title):
        """Crea un frame sezione con titolo"""
        section = tk.Frame(parent, bg='#f0f0f0')
        section.pack(fill='x', pady=10)
        
        # Titolo
        title_frame = tk.Frame(section, bg='#34495e')
        title_frame.pack(fill='x')
        
        title_label = tk.Label(title_frame, text=title, font=('Arial', 12, 'bold'), 
                              fg='white', bg='#34495e')
        title_label.pack(pady=10)
        
        # Content frame
        content_frame = tk.Frame(section, bg='white', relief='ridge', bd=1)
        content_frame.pack(fill='x')
        
        return content_frame
    
    def update_baselines(self):
        """Aggiorna la visualizzazione delle baseline"""
        # Pulisci container esistente
        for widget in self.baseline_container.winfo_children():
            widget.destroy()
            
        # Inizializza baseline_data se necessario
        current_count = self.baseline_count.get()
        while len(self.baseline_data) < current_count:
            self.baseline_data.append({
                'name': f'Baseline {len(self.baseline_data) + 1}',
                'quarters': 4,
                'description': '',
                'risk_level': 'Medio'
            })
        
        # Rimuovi baseline in eccesso
        self.baseline_data = self.baseline_data[:current_count]
        
        # Crea UI per ogni baseline
        for i in range(current_count):
            baseline_frame = tk.LabelFrame(self.baseline_container, text=f"Baseline {i+1}", 
                                         bg='white', font=('Arial', 10, 'bold'))
            baseline_frame.pack(fill='x', pady=5)
            
            # Grid per i campi
            grid = tk.Frame(baseline_frame, bg='white')
            grid.pack(fill='x', padx=10, pady=10)
            
            # Nome baseline
            tk.Label(grid, text="Nome:", bg='white', font=('Arial', 9)).grid(row=0, column=0, sticky='w')
            name_var = tk.StringVar(value=self.baseline_data[i]['name'])
            name_entry = tk.Entry(grid, textvariable=name_var, width=20, font=('Arial', 9))
            name_entry.grid(row=0, column=1, padx=5)
            name_entry.bind('<KeyRelease>', lambda e, idx=i, var=name_var: self.update_baseline_data(idx, 'name', var.get()))
            
            # Durata in quarters
            tk.Label(grid, text="Durata (Q):", bg='white', font=('Arial', 9)).grid(row=0, column=2, sticky='w', padx=(20, 0))
            quarters_var = tk.IntVar(value=self.baseline_data[i]['quarters'])
            quarters_spin = tk.Spinbox(grid, from_=2, to=12, textvariable=quarters_var, width=5, font=('Arial', 9))
            quarters_spin.grid(row=0, column=3, padx=5)
            quarters_spin.bind('<KeyRelease>', lambda e, idx=i, var=quarters_var: self.update_baseline_data(idx, 'quarters', var.get()))
            
            # Livello di rischio
            tk.Label(grid, text="Rischio:", bg='white', font=('Arial', 9)).grid(row=0, column=4, sticky='w', padx=(20, 0))
            risk_var = tk.StringVar(value=self.baseline_data[i]['risk_level'])
            risk_combo = ttk.Combobox(grid, textvariable=risk_var, values=['Basso', 'Medio', 'Alto', 'Molto Alto'], 
                                    width=10, font=('Arial', 9), state='readonly')
            risk_combo.grid(row=0, column=5, padx=5)
            risk_combo.bind('<<ComboboxSelected>>', lambda e, idx=i, var=risk_var: self.update_baseline_data(idx, 'risk_level', var.get()))
            
            # Descrizione
            tk.Label(grid, text="Descrizione:", bg='white', font=('Arial', 9)).grid(row=1, column=0, sticky='nw', pady=(10, 0))
            desc_text = tk.Text(grid, height=2, width=60, font=('Arial', 9))
            desc_text.grid(row=1, column=1, columnspan=5, sticky='ew', pady=(10, 0), padx=5)
            desc_text.insert('1.0', self.baseline_data[i]['description'])
            desc_text.bind('<KeyRelease>', lambda e, idx=i, text=desc_text: self.update_baseline_data(idx, 'description', text.get('1.0', 'end-1c')))
            
            grid.columnconfigure(1, weight=1)
    
    def update_baseline_data(self, index, field, value):
        """Aggiorna i dati di una baseline"""
        if index < len(self.baseline_data):
            self.baseline_data[index][field] = value
    
    def on_architecture_change(self):
        """Callback per cambio architettura"""
        # Qui puoi aggiungere logica per aggiornare preview o altro
        pass
    
    def preview_data(self):
        """Mostra anteprima dei dati configurati"""
        preview_window = tk.Toplevel(self.root)
        preview_window.title("📋 Anteprima Configurazione")
        preview_window.geometry("600x500")
        preview_window.configure(bg='white')
        
        # Testo con scroll
        text_frame = tk.Frame(preview_window, bg='white')
        text_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        # Contenuto anteprima
        preview_content = f"""
🚀 ANTEPRIMA CONFIGURAZIONE PROGETTO
====================================

📋 INFORMAZIONI PROGETTO:
• Nome: {self.project_name.get()}
• Cliente: {self.client_name.get() or 'Non specificato'}
• Descrizione: {self.project_description.get() or 'Non specificata'}

📊 BASELINE CONFIGURATE: {self.baseline_count.get()}
"""
        
        for i, baseline in enumerate(self.baseline_data, 1):
            preview_content += f"""
  Baseline {i}:
  • Nome: {baseline['name']}
  • Durata: {baseline['quarters']} quarters
  • Rischio: {baseline['risk_level']}
  • Descrizione: {baseline['description'] or 'Non specificata'}
"""
        
        arch_data = self.architectures[self.selected_architecture.get()]
        preview_content += f"""
🏗️ ARCHITETTURA SELEZIONATA:
• {arch_data['name']}
• {arch_data['description']}

📋 VOCI DI PROGETTO:
"""
        for i, item in enumerate(arch_data['items'], 1):
            preview_content += f"  {i}. {item}\n"
        
        preview_content += f"""
💰 TARIFFE CONFIGURATE:
"""
        for role, rate_var in self.rate_vars.items():
            preview_content += f"  • {role}: €{rate_var.get()}/giorno\n"
        
        text_widget.insert('1.0', preview_content)
        text_widget.config(state='disabled')
        
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
    
    def save_configuration(self):
        """Salva la configurazione in un file JSON"""
        config = {
            'project_name': self.project_name.get(),
            'client_name': self.client_name.get(),
            'project_description': self.project_description.get(),
            'baseline_count': self.baseline_count.get(),
            'selected_architecture': self.selected_architecture.get(),
            'baseline_data': self.baseline_data,
            'rates': {role: var.get() for role, var in self.rate_vars.items()}
        }
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Salva Configurazione"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Successo", "Configurazione salvata con successo!")
            except Exception as e:
                messagebox.showerror("Errore", f"Errore durante il salvataggio: {str(e)}")
    
    def load_configuration(self):
        """Carica una configurazione da file JSON"""
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Carica Configurazione"
        )
        
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # Aggiorna le variabili
                self.project_name.set(config.get('project_name', ''))
                self.client_name.set(config.get('client_name', ''))
                self.project_description.set(config.get('project_description', ''))
                self.baseline_count.set(config.get('baseline_count', 3))
                self.selected_architecture.set(config.get('selected_architecture', 'enterprise'))
                self.baseline_data = config.get('baseline_data', [])
                
                # Aggiorna tariffe
                rates_config = config.get('rates', {})
                for role, var in self.rate_vars.items():
                    var.set(rates_config.get(role, self.rates[role]))
                
                # Ricarica UI
                self.update_baselines()
                
                messagebox.showinfo("Successo", "Configurazione caricata con successo!")
                
            except Exception as e:
                messagebox.showerror("Errore", f"Errore durante il caricamento: {str(e)}")
    
    def generate_excel(self):
        """Genera il file Excel completo"""
        if not self.project_name.get().strip():
            messagebox.showerror("Errore", "Il nome del progetto è obbligatorio!")
            return
        
        # Chiedi dove salvare il file
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salva Template Excel",
            initialname=f"Quotazione_{self.project_name.get().replace(' ', '_')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Crea il workbook
            wb = openpyxl.Workbook()
            
            # Rimuovi il foglio default
            wb.remove(wb.active)
            
            # Crea i fogli
            self.create_dashboard_sheet(wb)
            self.create_configuration_sheet(wb)
            
            # Crea fogli baseline
            for i, baseline in enumerate(self.baseline_data):
                self.create_baseline_sheet(wb, i, baseline)
            
            self.create_quotation_sheet(wb)
            self.create_charts_sheet(wb)
            
            # Salva il file
            wb.save(filename)
            
            messagebox.showinfo("Successo!", 
                              f"Template Excel generato con successo!\n\nFile salvato: {filename}\n\n"
                              f"Il template include:\n"
                              f"• Dashboard esecutiva\n"
                              f"• {len(self.baseline_data)} baseline configurate\n"
                              f"• Calcoli automatici\n"
                              f"• Grafici dinamici\n"
                              f"• Formule Excel avanzate")
            
        except Exception as e:
            messagebox.showerror("Errore", f"Errore durante la generazione del file Excel:\n{str(e)}")
    
    def create_dashboard_sheet(self, wb):
        """Crea il foglio Dashboard"""
        ws = wb.create_sheet("📊 Dashboard", 0)
        
        # Stili
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        title_font = Font(bold=True, size=16)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Titolo
        ws['A1'] = f"🚀 DASHBOARD PROGETTO: {self.project_name.get()}"
        ws['A1'].font = Font(bold=True, size=18)
        ws.merge_cells('A1:F1')
        
        # Informazioni progetto
        row = 3
        ws[f'A{row}'] = "📋 INFORMAZIONI PROGETTO"
        ws[f'A{row}'].font = title_font
        
        row += 1
        ws[f'A{row}'] = "Cliente:"
        ws[f'B{row}'] = self.client_name.get() or "Non specificato"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Data Creazione:"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Baseline Configurate:"
        ws[f'B{row}'] = len(self.baseline_data)
        ws[f'A{row}'].font = Font(bold=True)
        
        # Tabella confronto baseline
        row += 3
        ws[f'A{row}'] = "📊 CONFRONTO BASELINE"
        ws[f'A{row}'].font = title_font
        
        row += 1
        headers = ['Baseline', 'Durata (Q)', 'Effort Totale (gg)', 'Costo Stimato (€)', 'Livello Rischio']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Dati baseline
        for i, baseline in enumerate(self.baseline_data):
            row += 1
            effort_total = 120 + (i * 40)  # Formula esempio
            cost = effort_total * 600  # Tariffa media
            
            data = [
                baseline['name'],
                baseline['quarters'],
                effort_total,
                cost,
                baseline['risk_level']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 3:  # Costo
                    cell.number_format = '€#,##0'
                cell.alignment = Alignment(horizontal='center')
        
        # Formattazione colonne
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
    
    def create_configuration_sheet(self, wb):
        """Crea il foglio Configurazione"""
        ws = wb.create_sheet("⚙️ Configurazione")
        
        # Stili
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        
        # Titolo
        ws['A1'] = "⚙️ CONFIGURAZIONE PROGETTO"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        
        # Architettura selezionata
        ws[f'A{row}'] = "🏗️ ARCHITETTURA PROGETTO"
        ws[f'A{row}'].font = title_font
        row += 1
        
        arch_data = self.architectures[self.selected_architecture.get()]
        ws[f'A{row}'] = "Tipo:"
        ws[f'B{row}'] = arch_data['name']
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Descrizione:"
        ws[f'B{row}'] = arch_data['description']
        ws[f'A{row}'].font = header_font
        row += 2
        
        # Voci di progetto
        ws[f'A{row}'] = "📋 VOCI DI PROGETTO"
        ws[f'A{row}'].font = title_font
        row += 1
        
        for i, item in enumerate(arch_data['items'], 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = item
            row += 1
        
        row += 1
        
        # Tariffe
        ws[f'A{row}'] = "💰 TARIFFE CONFIGURATE"
        ws[f'A{row}'].font = title_font
        row += 1
        
        ws[f'A{row}'] = "Ruolo"
        ws[f'B{row}'] = "Tariffa/giorno (€)"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1
        
        for role, rate_var in self.rate_vars.items():
            ws[f'A{row}'] = role
            ws[f'B{row}'] = rate_var.get()
            ws[f'B{row}'].number_format = '€#,##0'
            row += 1
        
        # Formattazione colonne
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def create_baseline_sheet(self, wb, index, baseline):
        """Crea un foglio per una baseline specifica"""
        ws = wb.create_sheet(f"📈 {baseline['name']}")
        
        # Stili
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Titolo
        ws['A1'] = f"📈 {baseline['name']} - DETTAGLIO EFFORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Informazioni baseline
        ws['A3'] = f"Durata: {baseline['quarters']} quarters"
        ws['A4'] = f"Livello Rischio: {baseline['risk_level']}"
        ws['A5'] = f"Descrizione: {baseline['description']}"
        
        # Headers tabella
        row = 7
        headers = ['Voce di Progetto']
        
        # Aggiungi colonne per ogni quarter
        for q in range(1, baseline['quarters'] + 1):
            headers.append(f'Q{q} (gg)')
        
        headers.extend(['Totale (gg)', 'Tariffa Media (€)', 'Costo Totale (€)'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Dati progetto
        arch_data = self.architectures[self.selected_architecture.get()]
        project_items = arch_data['items']
        
        total_effort = 0
        total_cost = 0
        
        for item_index, item in enumerate(project_items):
            row += 1
            ws.cell(row=row, column=1, value=item).border = border
            
            # Calcola effort per quarter (simulato)
            item_total_effort = 0
            quarters_effort = []
            
            base_effort = 5 + (index * 2)  # Effort base crescente per baseline
            for q in range(baseline['quarters']):
                # Distribuzione effort simulata
                if q == 0:  # Q1 - più effort per analisi/setup
                    effort = base_effort + (item_index % 3) * 2
                elif q == baseline['quarters'] - 1:  # Ultimo quarter - meno effort
                    effort = max(2, base_effort - 2)
                else:  # Quarter intermedi
                    effort = base_effort + (item_index % 4)
                
                quarters_effort.append(effort)
                item_total_effort += effort
                
                # Inserisci valore nella cella
                cell = ws.cell(row=row, column=q+2, value=effort)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Totale effort per item
            cell = ws.cell(row=row, column=baseline['quarters']+2, value=item_total_effort)
            cell.border = border
            cell.font = total_font
            cell.alignment = Alignment(horizontal='center')
            
            # Tariffa media (simulata in base al tipo di lavoro)
            avg_rate = 600 + (item_index % 3) * 100  # Tariffa variabile
            cell = ws.cell(row=row, column=baseline['quarters']+3, value=avg_rate)
            cell.border = border
            cell.number_format = '€#,##0'
            cell.alignment = Alignment(horizontal='center')
            
            # Costo totale
            item_cost = item_total_effort * avg_rate
            cell = ws.cell(row=row, column=baseline['quarters']+4, value=item_cost)
            cell.border = border
            cell.number_format = '€#,##0'
            cell.alignment = Alignment(horizontal='center')
            
            total_effort += item_total_effort
            total_cost += item_cost
        
        # Riga totali
        row += 1
        cell = ws.cell(row=row, column=1, value="TOTALE")
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        
        # Totali per quarter
        for q in range(baseline['quarters']):
            # Formula per sommare la colonna
            start_row = 8
            end_row = row - 1
            col_letter = get_column_letter(q + 2)
            formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            
            cell = ws.cell(row=row, column=q+2, value=formula)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Totale effort
        cell = ws.cell(row=row, column=baseline['quarters']+2, value=total_effort)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # Media tariffe
        avg_total_rate = total_cost / total_effort if total_effort > 0 else 0
        cell = ws.cell(row=row, column=baseline['quarters']+3, value=avg_total_rate)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '€#,##0'
        cell.alignment = Alignment(horizontal='center')
        
        # Totale costo
        cell = ws.cell(row=row, column=baseline['quarters']+4, value=total_cost)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '€#,##0'
        cell.alignment = Alignment(horizontal='center')
        
        # Formattazione colonne
        ws.column_dimensions['A'].width = 25
        for q in range(baseline['quarters']):
            ws.column_dimensions[get_column_letter(q+2)].width = 12
        ws.column_dimensions[get_column_letter(baseline['quarters']+2)].width = 15
        ws.column_dimensions[get_column_letter(baseline['quarters']+3)].width = 18
        ws.column_dimensions[get_column_letter(baseline['quarters']+4)].width = 18
    
    def create_quotation_sheet(self, wb):
        """Crea il foglio Quotazione finale"""
        ws = wb.create_sheet("💰 Quotazione")
        
        # Stili
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        title_font = Font(bold=True, size=16)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Titolo
        ws['A1'] = f"💰 QUOTAZIONE FINALE - {self.project_name.get()}"
        ws['A1'].font = Font(bold=True, size=18)
        ws.merge_cells('A1:F1')
        
        # Informazioni cliente
        row = 3
        ws[f'A{row}'] = "Cliente:"
        ws[f'B{row}'] = self.client_name.get() or "Non specificato"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Data Quotazione:"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Validità Offerta:"
        ws[f'B{row}'] = "30 giorni"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Tabella opzioni
        row += 3
        ws[f'A{row}'] = "🎯 OPZIONI DI PROGETTO"
        ws[f'A{row}'].font = title_font
        
        row += 1
        headers = ['Opzione', 'Descrizione', 'Durata', 'Effort (gg)', 'Costo Base (€)', 'Margine %', 'Prezzo Finale (€)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Opzioni baseline
        for i, baseline in enumerate(self.baseline_data):
            row += 1
            effort = 120 + (i * 40)
            base_cost = effort * 600
            margin = 20 + (i * 5)  # Margine crescente
            final_price = base_cost * (1 + margin/100)
            
            data = [
                baseline['name'],
                baseline['description'] or 'Scenario standard',
                f"{baseline['quarters']} quarters",
                effort,
                base_cost,
                f"{margin}%",
                final_price
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 7]:  # Colonne costi
                    cell.number_format = '€#,##0'
                cell.alignment = Alignment(horizontal='center')
        
        # Termini e condizioni
        row += 3
        ws[f'A{row}'] = "📋 TERMINI E CONDIZIONI"
        ws[f'A{row}'].font = title_font
        
        terms = [
            "• Prezzi espressi in Euro, IVA esclusa",
            "• Pagamenti: 30% anticipo, 40% SAL intermedio, 30% a consegna",
            "• Validità offerta: 30 giorni dalla data di emissione", 
            "• Eventuali modifiche ai requisiti comporteranno rinegoziazione",
            "• Include training del team cliente (8 ore)",
            "• Supporto post-go-live: 3 mesi inclusi",
            "• Documentazione tecnica completa inclusa"
        ]
        
        for term in terms:
            row += 1
            ws[f'A{row}'] = term
            ws.merge_cells(f'A{row}:G{row}')
        
        # Formattazione colonne
        column_widths = [15, 30, 12, 12, 15, 10, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_charts_sheet(self, wb):
        """Crea il foglio con i grafici"""
        ws = wb.create_sheet("📊 Grafici")
        
        # Titolo
        ws['A1'] = "📊 ANALISI GRAFICHE BASELINE"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Crea grafico comparativo baseline
        row = 3
        ws[f'A{row}'] = "Baseline"
        ws[f'B{row}'] = "Effort Totale (gg)"
        ws[f'C{row}'] = "Costo (€)"
        ws[f'D{row}'] = "Durata (Q)"
        
        # Dati per il grafico
        for i, baseline in enumerate(self.baseline_data):
            row += 1
            effort = 120 + (i * 40)
            cost = effort * 600
            
            ws[f'A{row}'] = baseline['name']
            ws[f'B{row}'] = effort
            ws[f'C{row}'] = cost
            ws[f'D{row}'] = baseline['quarters']
        
        # Crea grafico a colonne
        chart = ColumnChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Confronto Effort per Baseline"
        chart.y_axis.title = 'Giorni di Effort'
        chart.x_axis.title = 'Baseline'
        
        # Dati per il grafico
        data = Reference(ws, min_col=2, min_row=3, max_row=3+len(self.baseline_data), max_col=2)
        categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(self.baseline_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Posiziona il grafico
        ws.add_chart(chart, "F3")
        
        # Secondo grafico per i costi
        chart2 = ColumnChart()
        chart2.type = "col"
        chart2.style = 12
        chart2.title = "Confronto Costi per Baseline"
        chart2.y_axis.title = 'Costo (€)'
        chart2.x_axis.title = 'Baseline'
        
        data2 = Reference(ws, min_col=3, min_row=3, max_row=3+len(self.baseline_data), max_col=3)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(categories)
        
        ws.add_chart(chart2, "F18")
        
        # Formattazione
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12

def main():
    """Funzione principale"""
    root = tk.Tk()
    app = QuotationGenerator(root)
    
    # Icona e configurazioni finali
    try:
        # Se hai un'icona, puoi aggiungerla qui
        # root.iconbitmap('icon.ico')
        pass
    except:
        pass
    
    # Centra la finestra
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()
